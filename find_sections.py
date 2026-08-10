import openpyxl

file_path = "/Users/suphakorn/EA Market Place/MT5 EA Reports/ReportHistory-97053088-EasyM Max-Apr27-Aug08.xlsx"

wb = openpyxl.load_workbook(file_path, read_only=True)
sheet = wb.active

for r in range(1, sheet.max_row + 1):
    val = sheet.cell(row=r, column=1).value
    if val is not None and not isinstance(val, (int, float)) and len(str(val)) < 50:
        # Check if the row contains any section headers
        if any(h in str(val) for h in ["Positions", "Open Positions", "Orders", "Deals", "Summary"]):
            print(f"Row {r}: '{val}'")
